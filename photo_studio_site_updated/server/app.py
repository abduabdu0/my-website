from flask import Flask, request, jsonify, send_from_directory
from openpyxl import Workbook, load_workbook
from datetime import datetime
import os, smtplib
from email.message import EmailMessage

# --- Config: update these values before running ---
SMTP_HOST = 'smtp.example.com'
SMTP_PORT = 587
SMTP_USER = 'you@example.com'
SMTP_PASS = 'yourpassword'
EMAIL_TO = 'your_destination@example.com'
# -----------------------------------------------

DATA_FILE = 'submissions.xlsx'

app = Flask(__name__, static_folder='.', static_url_path='')

def ensure_workbook(path):
    if not os.path.exists(path):
        wb = Workbook()
        ws = wb.active
        ws.append(['Дата', 'Имя', 'Контакт', 'Заметки'])
        wb.save(path)

def append_submission(path, name, phone, note):
    ensure_workbook(path)
    wb = load_workbook(path)
    ws = wb.active
    ws.append([datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S'), name, phone, note])
    wb.save(path)

def send_email_with_attachment(smtp_host, smtp_port, smtp_user, smtp_pass, to_email, subject, body, attachment_path):
    msg = EmailMessage()
    msg['Subject'] = subject
    msg['From'] = smtp_user
    msg['To'] = to_email
    msg.set_content(body)
    with open(attachment_path,'rb') as f:
        data = f.read()
    msg.add_attachment(data, maintype='application', subtype='vnd.openxmlformats-officedocument.spreadsheetml.sheet', filename=os.path.basename(attachment_path))
    with smtplib.SMTP(smtp_host, smtp_port) as s:
        s.starttls()
        s.login(smtp_user, smtp_pass)
        s.send_message(msg)

@app.route('/submit', methods=['POST'])
def submit():
    try:
        data = request.get_json()
        name = data.get('name','')
        phone = data.get('phone','')
        note = data.get('note','')
        append_submission(DATA_FILE, name, phone, note)
        # send email with attachment
        try:
            send_email_with_attachment(SMTP_HOST, SMTP_PORT, SMTP_USER, SMTP_PASS, EMAIL_TO, 'Новая заявка StudioPro', f'Новая заявка от {name} ({phone})', DATA_FILE)
        except Exception as e:
            # don't fail completely — just report email error
            return jsonify({'ok': True, 'warning': f'Сохранено, но не удалось отправить email: {e}'}), 200
        return jsonify({'ok': True}), 200
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500

# Static file serving (for local testing)
@app.route('/', defaults={'path':'index.html'})
@app.route('/<path:path>')
def static_proxy(path):
    return send_from_directory('.', path)

if __name__ == '__main__':
    print('Перед запуском отредактируйте конфигурацию SMTP в server/app.py')
    app.run(host='0.0.0.0', port=5000, debug=True)
